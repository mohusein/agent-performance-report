import pandas as pd
import warnings
warnings.filterwarnings("ignore", category=FutureWarning)
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Helper: convert "H:MM:SS" time string to total seconds
def time_to_seconds(t):
    try:
        parts = str(t).strip().split(':')
        if len(parts) == 3:
            return int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])
        elif len(parts) == 2:
            return int(parts[0]) * 60 + int(parts[1])
    except:
        pass
    return 0

# 1. Load the Ytel CSV — auto-detect format
try:
    # Try reading with headers on first row (new download format)
    df_test = pd.read_csv("ytel_agent_report.csv", nrows=2)
    if 'full_name' in df_test.columns or 'calls' in df_test.columns:
        # New format: headers already on row 0, lowercase
        df = pd.read_csv("ytel_agent_report.csv")
        df.columns = df.columns.str.strip()
        # Normalize column names to match expected format
        col_map = {
            'full_name': 'Name', 'id': 'Agent Id', 'user_group': 'User group',
            'calls': 'CALLS', 'time': 'TIME', 'pause': 'PAUSE', 'pause_avg': 'PAUSE AVG',
            'wait': 'WAIT', 'wait_avg': 'WAITAVG', 'talk': 'TALK', 'talk_avg': 'TALK AVG',
            'dispo': 'DISPO', 'dispo_avg': 'DISPO AVG', 'dead': 'DEAD', 'dead_avg': 'DEAD AVG',
            'customer': 'CUSTOMER', 'customer_avg': 'CUST AVG',
            'AMLVM': 'AMLVM', 'AMNAVM': 'AMNAVM', 'AMNVM': 'AMNVM',
            'CALLBK': 'CALLBK', 'DNC': 'DNC', 'NI': 'NI',
        }
        df.rename(columns={k: v for k, v in col_map.items() if k in df.columns}, inplace=True)
        df.dropna(how='all', inplace=True)
        # Use 'Agent Id' or 'id' whichever exists after renaming
        id_col = 'Agent Id' if 'Agent Id' in df.columns else 'id'
        df = df[pd.to_numeric(df[id_col], errors='coerce').notna()].copy()
        if id_col != 'Agent Id':
            df.rename(columns={id_col: 'Agent Id'}, inplace=True)
    else:
        # Old format: report title on row 0, real headers on row 4
        df = pd.read_csv("ytel_agent_report.csv", header=4)
        df.columns = df.columns.str.strip()
        df.dropna(how='all', inplace=True)
        df = df[pd.to_numeric(df['Agent Id'], errors='coerce').notna()].copy()
except FileNotFoundError:
    print("Error: ytel_agent_report.csv not found.")
    exit(1)

print("Columns found:", df.columns.tolist())
print(f"Rows loaded: {len(df)}")

# 2. Convert time columns to seconds for math
for col in ['TIME', 'TALK', 'PAUSE', 'WAIT']:
    if col in df.columns:
        df[col + '_sec'] = df[col].apply(time_to_seconds)

# 3. Calculate percentage columns
df['CALLS'] = pd.to_numeric(df['CALLS'], errors='coerce').fillna(0)

if 'TIME_sec' in df.columns and df['TIME_sec'].sum() > 0:
    total = df['TIME_sec'].replace(0, pd.NA)
    df['Talk %']  = (df['TALK_sec']  / total).fillna(0)
    df['Pause %'] = (df['PAUSE_sec'] / total).fillna(0)
    df['Wait %']  = (df['WAIT_sec']  / total).fillna(0)

# Convert DISPO, DEAD, CUSTOMER avg columns to seconds then percentages
for col in ['DISPO AVG', 'DEAD AVG', 'CUST AVG']:
    if col in df.columns:
        df[col + '_sec'] = df[col].apply(time_to_seconds)

if 'TIME_sec' in df.columns and df['TIME_sec'].sum() > 0:
    total = df['TIME_sec'].replace(0, pd.NA)
    if 'DISPO AVG_sec' in df.columns:
        df['Dispo Avg %'] = (df['DISPO AVG_sec'] / total).fillna(0)
    if 'DEAD AVG_sec' in df.columns:
        df['Dead Avg %']  = (df['DEAD AVG_sec']  / total).fillna(0)
    if 'CUST AVG_sec' in df.columns:
        df['Cust Avg %']  = (df['CUST AVG_sec']  / total).fillna(0)

# 4. Conversion % = AMNVM / CALLS
sale_col = next((c for c in ['SALE', 'AMNVM'] if c in df.columns), None)
if sale_col:
    df[sale_col] = pd.to_numeric(df[sale_col], errors='coerce').fillna(0)
    df['Conversion %'] = (df[sale_col] / df['CALLS'].replace(0, pd.NA)).fillna(0)
    print(f"Conversion % calculated using '{sale_col}' / 'CALLS'")
else:
    df['Conversion %'] = 0
    print("Warning: Could not find sale column. Conversion % set to 0.")

# 5. Build report dataframe
report_columns = [
    'Agent Id', 'Name', 'User group', 'CALLS',
    'Talk %', 'Pause %', 'Wait %',
    'Dispo Avg %', 'Dead Avg %', 'Cust Avg %',
    'AMNVM', 'AMLVM', 'AMNAVM', 'CALLBK', 'NI', 'DNC',
    'Conversion %'
]
report_columns = [col for col in report_columns if col in df.columns]
df_report = df[report_columns].copy()

rename_map = {
    'Agent Id':   'Agent ID',
    'Name':       'Agent Name',
    'User group': 'Team',
    'CALLS':      'Total Calls',
    'AMNVM':      'Sales',
    'CALLBK':     'Callbacks',
    'NI':         'Not Interested',
    'DNC':        'DNC',
}
df_report.rename(columns={k: v for k, v in rename_map.items() if k in df_report.columns}, inplace=True)

# Columns that should be formatted as percentages
pct_columns = {'Talk %', 'Pause %', 'Wait %', 'Dispo Avg %', 'Dead Avg %', 'Cust Avg %', 'Conversion %'}

# 6. Build Excel workbook
wb = Workbook()
ws = wb.active
ws.title = "Agent Performance"

for r in dataframe_to_rows(df_report, index=False, header=True):
    ws.append(r)

# 7. Style header row
header_fill = PatternFill(start_color="2F75B6", end_color="2F75B6", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

# 8. Apply percentage number format to all pct columns
for col_idx, col_name in enumerate(df_report.columns, start=1):
    if col_name in pct_columns:
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col_idx).number_format = '0.0%'

# 9. Color-code rows by Conversion % — green >= 10%, red < 10%
#    Green rows: white text on green fill
#    Red rows:   leave as-is (no text color change)
red_fill   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
white_font = Font(color="FFFFFF")

if 'Conversion %' in df_report.columns:
    conv_col_idx = list(df_report.columns).index('Conversion %') + 1

    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=conv_col_idx)

        if cell.value is not None:
            if cell.value >= 0.10:
                # White row with default text
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = white_fill
            else:
                # Red row — only apply fill, leave text alone
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = red_fill

# 10. Add totals row
total_fill = PatternFill(start_color="2F75B6", end_color="2F75B6", fill_type="solid")
total_font = Font(color="FFFFFF", bold=True)

totals_row = []
for col_name in df_report.columns:
    if col_name == 'Agent ID':
        totals_row.append('TOTAL')
    elif col_name in ('Agent Name', 'Team'):
        totals_row.append('')
    elif col_name in pct_columns:
        # Average of percentage columns
        totals_row.append(df_report[col_name].mean())
    else:
        # Sum numeric columns
        numeric = pd.to_numeric(df_report[col_name], errors='coerce')
        totals_row.append(numeric.sum() if numeric.notna().any() else '')

ws.append(totals_row)
total_row_idx = ws.max_row

# Style the totals row — blue background, white bold text
for col_idx, col_name in enumerate(df_report.columns, start=1):
    cell = ws.cell(row=total_row_idx, column=col_idx)
    cell.fill = total_fill
    cell.font = total_font
    cell.alignment = Alignment(horizontal="center")
    if col_name in pct_columns:
        cell.number_format = '0.0%'

# 11. Auto-fit column widths
for col in ws.columns:
    max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
    ws.column_dimensions[col[0].column_letter].width = max_len + 4

# 12. Save Excel
wb.save("Formatted_Agent_Report.xlsx")
print("Excel report generated: Formatted_Agent_Report.xlsx")
